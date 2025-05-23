import openpyxl
from openpyxl.styles import Border, Side, Alignment
import os
"""
入口，将需要计算的文件放入input文件夹中运行。
结果保存在output中。
"""
folder_path = 'input'
result_folder = 'output'

class Conversion_Parameters:
    def __init__(self):
        """
        MU_TO_HECTARES:  亩转公顷系数

        LOSS_CROP_AN: 农作物氨氮
        LOSS_CROP_TN： 农作物总氮
        LOSS_CROP_TP： 农作物总磷
        LOSS_ORCHARD_AN： 园地氨氮
        LOSS_ORCHARD_TN： 园地总氮
        LOSS_ORCHARD_TP： 园地总磷

        CF_FOOD_MAX: 粮食最大复种指数
        CF_FOOD_MIN: 粮食最小复种指数

        CF_OIL_MAX： 油料最大复种指数
        CF_OIL_MIN： 油料最小复种指数

        CF_MEDICINAL_MAX： 中草药最大复种指数
        CF_MEDICINAL_MIN： 中草药最小复种指数

        CF_VEGETABLE_FACILITY_MAX： 蔬菜设施最大复种指数
        CF_VEGETABLE_FACILITY_MIN： 蔬菜设施最小复种指数
        CF_VEGETABLE_OPEN_MAX： 蔬菜露天最大复种指数
        CF_VEGETABLE_OPEN_MIN： 蔬菜露天最小复种指数

        CF_MELON_FACILITY_MAX: 瓜果设施最大复种指数
        CF_MELON_FACILITY_MIN: 瓜果设施最小复种指数
        CF_MELON_OPEN_MAX: 瓜果露天最大复种指数
        CF_MELON_OPEN_MIN: 瓜果露天最小复种指数

        CF_TEA_MAX: 茶园最大复种指数
        CF_TEA_MIN: 茶园最小复种指数

        CF_FLOWER_FACILITY_MAX： 花卉设施最大复种指数
        CF_FLOWER_FACILITY_MIN： 花卉设施最小复种指数
        CF_FLOWER_OPEN_MAX： 花卉露天最大复种指数
        CF_FLOWER_OPEN_MIN： 花卉露天最小复种指数

        CF_FRUIT_MAX： 林果园最大复种指数
        CF_FRUIT_MIN： 林果园最小复种指数

        CF_MULBERRY_MAX： 桑园最大复种指数
        CF_MULBERRY_MIN： 桑园最小复种指数

        CF_SEEDING_MAX： 苗木最大复种指数
        CF_SEEDING_MIN： 苗木最小复种指数

        CF_OTHER_MAX： 其他最大复种指数
        CF_OTHER_MIN： 其他最小复种指数
        """
        self.MU_TO_HECTARES = 0.0666666666667
        self.LOSS_CROP_AN = 0.431
        self.LOSS_CROP_TN = 6.387
        self.LOSS_CROP_TP = 0.509
        self.LOSS_ORCHARD_AN = 0.205
        self.LOSS_ORCHARD_TN = 3.087
        self.LOSS_ORCHARD_TP = 0.335

        self.CF_FOOD_MAX = 1.2
        self.CF_FOOD_MIN = 1.2

        self.CF_OIL_MAX = 2
        self.CF_OIL_MIN = 1

        self.CF_MEDICINAL_MAX = 2
        self.CF_MEDICINAL_MIN = 1

        self.CF_VEGETABLE_FACILITY_MAX = 5
        self.CF_VEGETABLE_FACILITY_MIN = 5
        self.CF_VEGETABLE_OPEN_MAX = 4
        self.CF_VEGETABLE_OPEN_MIN = 2

        self.CF_MELON_FACILITY_MAX = 5
        self.CF_MELON_FACILITY_MIN = 5
        self.CF_MELON_OPEN_MAX = 4
        self.CF_MELON_OPEN_MIN = 2

        self.CF_TEA_MAX = 1
        self.CF_TEA_MIN = 1

        self.CF_FLOWER_FACILITY_MAX = 5
        self.CF_FLOWER_FACILITY_MIN = 5
        self.CF_FLOWER_OPEN_MAX = 4
        self.CF_FLOWER_OPEN_MIN = 1

        self.CF_FRUIT_MAX = 5
        self.CF_FRUIT_MIN = 1.5

        self.CF_MULBERRY_MAX = 1
        self.CF_MULBERRY_MIN = 1

        self.CF_SEEDING_MAX = 1
        self.CF_SEEDING_MIN = 1

        self.CF_OTHER_MAX = 1
        self.CF_OTHER_MIN = 1

# 遍历需要计算的文件夹（input)中的所有文件并且创建我需要计算和转化单位的表
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        print(file_name)
        CP = Conversion_Parameters()
        wb = openpyxl.load_workbook(f"input/{file_name}", data_only = True)
        original_sheet = wb.active
        original_sheet.title = '原表（亩）'

        converted_sheet_name = '原表（公顷）'
        wb.create_sheet(title=converted_sheet_name)
        converted_sheet = wb[converted_sheet_name]   # 定位到转换种植面积为公顷后的表

        crop_area_sheet_name = '作物面积统计'
        wb.create_sheet(title=crop_area_sheet_name)
        crop_area_sheet = wb[crop_area_sheet_name]

        SUM_sheet_name = '总表'
        wb.create_sheet(title=SUM_sheet_name)
        SUM_sheet = wb[SUM_sheet_name]

        AN_MAX_sheet_name = '氨氮统计_MAX'
        wb.create_sheet(title=AN_MAX_sheet_name)
        AN_MAX_sheet = wb[AN_MAX_sheet_name]

        TN_MAX_sheet_name = '总氮统计_MAX'
        wb.create_sheet(title=TN_MAX_sheet_name)
        TN_MAX_sheet = wb[TN_MAX_sheet_name]

        TP_MAX_sheet_name = '总磷统计_MAX'
        wb.create_sheet(title=TP_MAX_sheet_name)
        TP_MAX_sheet = wb[TP_MAX_sheet_name]

        AN_MIN_sheet_name = '氨氮统计_MIN'
        wb.create_sheet(title=AN_MIN_sheet_name)
        AN_MIN_sheet = wb[AN_MIN_sheet_name]

        TN_MIN_sheet_name = '总氮统计_MIN'
        wb.create_sheet(title=TN_MIN_sheet_name)
        TN_MIN_sheet = wb[TN_MIN_sheet_name]

        TP_MIN_sheet_name = '总磷统计_MIN'
        wb.create_sheet(title=TP_MIN_sheet_name)
        TP_MIN_sheet = wb[TP_MIN_sheet_name]

        def Mu_to_hec():
            """
            转化为公顷并计算大类总面积
            """
            for row in original_sheet.iter_rows(min_row=4,max_row=39, min_col=1, max_col=original_sheet.max_column):  # 遍历需要计算数据
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        converted_value = round(cell.value * CP.MU_TO_HECTARES, 12)  #公顷转化为亩且保留12位精度
                        converted_sheet[cell.coordinate].value = converted_value
                    else:
                        converted_sheet[cell.coordinate].value = cell.value
            converted_sheet.delete_rows(1, 3)
            converted_sheet.delete_rows(2, 1)

        def Crop_area_statistics():
            """
            作物面积统计
            """
            headers = ['', '乡镇1', '乡镇2', '乡镇3','合计']
            for col, header in enumerate(headers, start=1):
                crop_area_sheet.cell(row=1, column=col, value=header)

            crop_area_sheet.cell(1,1,'乡镇')
            crop_area_sheet.cell(2,1,'粮食')
            crop_area_sheet.cell(3,1,'油料')
            crop_area_sheet.cell(4,1,'中草药材')
            crop_area_sheet.cell(5,1,'蔬菜')
            crop_area_sheet.cell(6,1,'瓜果')
            crop_area_sheet.cell(7,1,'茶园')
            crop_area_sheet.cell(8,1,'花卉')
            crop_area_sheet.cell(9,1,'林果园')
            crop_area_sheet.cell(10,1,'桑园')
            crop_area_sheet.cell(11,1,'苗木')
            crop_area_sheet.cell(12,1,'其他')

            start_col = 3

            for col in range(start_col, converted_sheet.max_column + 1):
                # 从公顷表的第3列第12行开始复制，从面积表的第2列第2行开始粘贴
                crop_area_sheet.cell(column=col - start_col + 2, row=2, value=converted_sheet.cell(row=12, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=3, value=converted_sheet.cell(row=14, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=4, value=converted_sheet.cell(row=16, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=5, value=converted_sheet.cell(row=19, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=6, value=converted_sheet.cell(row=22, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=7, value=converted_sheet.cell(row=24, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=8, value=converted_sheet.cell(row=27, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=9, value=converted_sheet.cell(row=29, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=10, value=converted_sheet.cell(row=31, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=11, value=converted_sheet.cell(row=33, column=col).value)
                crop_area_sheet.cell(column=col - start_col + 2, row=12, value=converted_sheet.cell(row=35, column=col).value)

        def AN_MAX_table_create():
            """
            氨氮——MAX表创建
            """
            headers = ['', '乡镇1', '乡镇2', '乡镇3','合计']
            for col, header in enumerate(headers, start=1):
                AN_MAX_sheet.cell(row=1, column=col, value=header)

            AN_MAX_sheet.cell(1, 1, '乡镇')
            AN_MAX_sheet.cell(2, 1, '粮食')
            AN_MAX_sheet.cell(3, 1, '油料')
            AN_MAX_sheet.cell(4, 1, '中草药材')
            AN_MAX_sheet.cell(5, 1, '蔬菜')
            AN_MAX_sheet.cell(6, 1, '瓜果')
            AN_MAX_sheet.cell(7, 1, '茶园')
            AN_MAX_sheet.cell(8, 1, '花卉')
            AN_MAX_sheet.cell(9, 1, '林果园')
            AN_MAX_sheet.cell(10, 1, '桑园')
            AN_MAX_sheet.cell(11, 1, '苗木')
            AN_MAX_sheet.cell(12, 1, '其他')

            start_col = 3
            for col in range(start_col, converted_sheet.max_column + 1):
                temp = CP.LOSS_CROP_AN * ((converted_sheet.cell(column=col, row=12).value or 0) / 1000)/(CP.CF_FOOD_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=2,value=temp)  # 粮食

                temp = CP.LOSS_CROP_AN * ((converted_sheet.cell(column=col, row=14).value or 0) / 1000)/(CP.CF_OIL_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=3,value=temp)  # 油料

                temp = CP.LOSS_CROP_AN * ((converted_sheet.cell(column=col, row=16).value or 0) / 1000)/(CP.CF_MEDICINAL_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=4,value=temp)  # 中药材

                temp = CP.LOSS_CROP_AN * (((converted_sheet.cell(column=col, row=17).value or 0)/1000)/(CP.CF_VEGETABLE_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=18).value or 0)/ 1000) / (CP.CF_VEGETABLE_OPEN_MAX))
                AN_MAX_sheet.cell(column=col - start_col + 2, row=5,value=temp)  # 蔬菜类

                temp = CP.LOSS_ORCHARD_AN * (((converted_sheet.cell(column=col, row=20).value or 0)/1000)/(CP.CF_MELON_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=21).value or 0)/ 1000) / (CP.CF_MELON_OPEN_MAX))
                AN_MAX_sheet.cell(column=col - start_col + 2, row=6,value=temp)  # 瓜果类

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=24).value or 0) / 1000)/(CP.CF_TEA_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=7,value=temp)  # 茶园

                temp = CP.LOSS_ORCHARD_AN * (((converted_sheet.cell(column=col, row=25).value or 0)/1000)/(CP.CF_FLOWER_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=26).value or 0)/ 1000) / (CP.CF_FLOWER_OPEN_MAX))
                AN_MAX_sheet.cell(column=col - start_col + 2, row=8,value=temp)  # 花卉类

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=29).value or 0) / 1000)/(CP.CF_FRUIT_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=9,value=temp)  # 林果园

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=31).value or 0) / 1000)/(CP.CF_MULBERRY_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=10,value=temp)  # 桑园

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=33).value or 0) / 1000)/(CP.CF_SEEDING_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=11,value=temp)  # 苗木

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=35).value or 0) / 1000)/(CP.CF_OTHER_MAX)
                AN_MAX_sheet.cell(column=col - start_col + 2, row=12,value=temp)  # 其他

        def AN_MIN_table_create():
            """
            氨氮——MIN表创建
            """
            headers = ['', '乡镇1', '乡镇2', '乡镇3','合计']
            for col, header in enumerate(headers, start=1):
                AN_MIN_sheet.cell(row=1, column=col, value=header)

            AN_MIN_sheet.cell(1, 1, '乡镇')
            AN_MIN_sheet.cell(2, 1, '粮食')
            AN_MIN_sheet.cell(3, 1, '油料')
            AN_MIN_sheet.cell(4, 1, '中草药材')
            AN_MIN_sheet.cell(5, 1, '蔬菜')
            AN_MIN_sheet.cell(6, 1, '瓜果')
            AN_MIN_sheet.cell(7, 1, '茶园')
            AN_MIN_sheet.cell(8, 1, '花卉')
            AN_MIN_sheet.cell(9, 1, '林果园')
            AN_MIN_sheet.cell(10, 1, '桑园')
            AN_MIN_sheet.cell(11, 1, '苗木')
            AN_MIN_sheet.cell(12, 1, '其他')

            start_col = 3
            for col in range(start_col, converted_sheet.max_column + 1):
                temp = CP.LOSS_CROP_AN * ((converted_sheet.cell(column=col, row=12).value or 0) / 1000)/(CP.CF_FOOD_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=2,value=temp)  # 粮食

                temp = CP.LOSS_CROP_AN * ((converted_sheet.cell(column=col, row=14).value or 0) / 1000)/(CP.CF_OIL_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=3,value=temp)  # 油料

                temp = CP.LOSS_CROP_AN * ((converted_sheet.cell(column=col, row=16).value or 0) / 1000)/(CP.CF_MEDICINAL_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=4,value=temp)  # 中药材

                temp = CP.LOSS_CROP_AN * (((converted_sheet.cell(column=col, row=17).value or 0)/1000)/(CP.CF_VEGETABLE_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=18).value or 0)/ 1000) / (CP.CF_VEGETABLE_OPEN_MIN))
                AN_MIN_sheet.cell(column=col - start_col + 2, row=5,value=temp)  # 蔬菜类

                temp = CP.LOSS_ORCHARD_AN * (((converted_sheet.cell(column=col, row=20).value or 0)/1000)/(CP.CF_MELON_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=21).value or 0)/ 1000) / (CP.CF_MELON_OPEN_MIN))
                AN_MIN_sheet.cell(column=col - start_col + 2, row=6,value=temp)  # 瓜果类

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=24).value or 0) / 1000)/(CP.CF_TEA_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=7,value=temp)  # 茶园

                temp = CP.LOSS_ORCHARD_AN * (((converted_sheet.cell(column=col, row=25).value or 0)/1000)/(CP.CF_FLOWER_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=26).value or 0)/ 1000) / (CP.CF_FLOWER_OPEN_MIN))
                AN_MIN_sheet.cell(column=col - start_col + 2, row=8,value=temp)  # 花卉类

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=29).value or 0) / 1000)/(CP.CF_FRUIT_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=9,value=temp)  # 林果园

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=31).value or 0) / 1000)/(CP.CF_MULBERRY_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=10,value=temp)  # 桑园

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=33).value or 0) / 1000)/(CP.CF_SEEDING_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=11,value=temp)  # 苗木

                temp = CP.LOSS_ORCHARD_AN * ((converted_sheet.cell(column=col, row=35).value or 0) / 1000)/(CP.CF_OTHER_MIN)
                AN_MIN_sheet.cell(column=col - start_col + 2, row=12,value=temp)  # 其他

        def TN_MAX_table_create():
            """
            总氮——MAX表创建
            """
            headers = ['', '乡镇1', '乡镇2', '乡镇3','合计']
            for col, header in enumerate(headers, start=1):
                TN_MAX_sheet.cell(row=1, column=col, value=header)

            TN_MAX_sheet.cell(1, 1, '乡镇')
            TN_MAX_sheet.cell(2, 1, '粮食')
            TN_MAX_sheet.cell(3, 1, '油料')
            TN_MAX_sheet.cell(4, 1, '中草药材')
            TN_MAX_sheet.cell(5, 1, '蔬菜')
            TN_MAX_sheet.cell(6, 1, '瓜果')
            TN_MAX_sheet.cell(7, 1, '茶园')
            TN_MAX_sheet.cell(8, 1, '花卉')
            TN_MAX_sheet.cell(9, 1, '林果园')
            TN_MAX_sheet.cell(10, 1, '桑园')
            TN_MAX_sheet.cell(11, 1, '苗木')
            TN_MAX_sheet.cell(12, 1, '其他')

            start_col = 3
            for col in range(start_col, converted_sheet.max_column + 1):
                temp = CP.LOSS_CROP_TN * ((converted_sheet.cell(column=col, row=12).value or 0) / 1000)/(CP.CF_FOOD_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=2,value=temp)  # 粮食

                temp = CP.LOSS_CROP_TN * ((converted_sheet.cell(column=col, row=14).value or 0) / 1000)/(CP.CF_OIL_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=3,value=temp)  # 油料

                temp = CP.LOSS_CROP_TN * ((converted_sheet.cell(column=col, row=16).value or 0) / 1000)/(CP.CF_MEDICINAL_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=4,value=temp)  # 中药材

                temp = CP.LOSS_CROP_TN * (((converted_sheet.cell(column=col, row=17).value or 0)/1000)/(CP.CF_VEGETABLE_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=18).value or 0)/ 1000) / (CP.CF_VEGETABLE_OPEN_MAX))
                TN_MAX_sheet.cell(column=col - start_col + 2, row=5,value=temp)  # 蔬菜类

                temp = CP.LOSS_ORCHARD_TN * (((converted_sheet.cell(column=col, row=20).value or 0)/1000)/(CP.CF_MELON_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=21).value or 0)/ 1000) / (CP.CF_MELON_OPEN_MAX))
                TN_MAX_sheet.cell(column=col - start_col + 2, row=6,value=temp)  # 瓜果类

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=24).value or 0) / 1000)/(CP.CF_TEA_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=7,value=temp)  # 茶园

                temp = CP.LOSS_ORCHARD_TN * (((converted_sheet.cell(column=col, row=25).value or 0)/1000)/(CP.CF_FLOWER_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=26).value or 0)/ 1000) / (CP.CF_FLOWER_OPEN_MAX))
                TN_MAX_sheet.cell(column=col - start_col + 2, row=8,value=temp)  # 花卉类

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=29).value or 0) / 1000)/(CP.CF_FRUIT_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=9,value=temp)  # 林果园

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=31).value or 0) / 1000)/(CP.CF_MULBERRY_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=10,value=temp)  # 桑园

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=33).value or 0) / 1000)/(CP.CF_SEEDING_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=11,value=temp)  # 苗木

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=35).value or 0) / 1000)/(CP.CF_OTHER_MAX)
                TN_MAX_sheet.cell(column=col - start_col + 2, row=12,value=temp)  # 其他

        def TN_MIN_table_create():
            """
            总氮——MIN表创建
            """
            headers = ['', '乡镇1', '乡镇2', '乡镇3','合计']
            for col, header in enumerate(headers, start=1):
                TN_MIN_sheet.cell(row=1, column=col, value=header)

            TN_MIN_sheet.cell(1, 1, '乡镇')
            TN_MIN_sheet.cell(2, 1, '粮食')
            TN_MIN_sheet.cell(3, 1, '油料')
            TN_MIN_sheet.cell(4, 1, '中草药材')
            TN_MIN_sheet.cell(5, 1, '蔬菜')
            TN_MIN_sheet.cell(6, 1, '瓜果')
            TN_MIN_sheet.cell(7, 1, '茶园')
            TN_MIN_sheet.cell(8, 1, '花卉')
            TN_MIN_sheet.cell(9, 1, '林果园')
            TN_MIN_sheet.cell(10, 1, '桑园')
            TN_MIN_sheet.cell(11, 1, '苗木')
            TN_MIN_sheet.cell(12, 1, '其他')

            start_col = 3
            for col in range(start_col, converted_sheet.max_column + 1):
                temp = CP.LOSS_CROP_TN * ((converted_sheet.cell(column=col, row=12).value or 0) / 1000)/(CP.CF_FOOD_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=2,value=temp)  # 粮食

                temp = CP.LOSS_CROP_TN * ((converted_sheet.cell(column=col, row=14).value or 0) / 1000)/(CP.CF_OIL_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=3,value=temp)  # 油料

                temp = CP.LOSS_CROP_TN * ((converted_sheet.cell(column=col, row=16).value or 0) / 1000)/(CP.CF_MEDICINAL_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=4,value=temp)  # 中药材

                temp = CP.LOSS_CROP_TN * (((converted_sheet.cell(column=col, row=17).value or 0)/1000)/(CP.CF_VEGETABLE_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=18).value or 0)/ 1000) / (CP.CF_VEGETABLE_OPEN_MIN))
                TN_MIN_sheet.cell(column=col - start_col + 2, row=5,value=temp)  # 蔬菜类

                temp = CP.LOSS_ORCHARD_TN * (((converted_sheet.cell(column=col, row=20).value or 0)/1000)/(CP.CF_MELON_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=21).value or 0)/ 1000) / (CP.CF_MELON_OPEN_MIN))
                TN_MIN_sheet.cell(column=col - start_col + 2, row=6,value=temp)  # 瓜果类

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=24).value or 0) / 1000)/(CP.CF_TEA_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=7,value=temp)  # 茶园

                temp = CP.LOSS_ORCHARD_TN * (((converted_sheet.cell(column=col, row=25).value or 0)/1000)/(CP.CF_FLOWER_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=26).value or 0)/ 1000) / (CP.CF_FLOWER_OPEN_MIN))
                TN_MIN_sheet.cell(column=col - start_col + 2, row=8,value=temp)  # 花卉类

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=29).value or 0) / 1000)/(CP.CF_FRUIT_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=9,value=temp)  # 林果园

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=31).value or 0) / 1000)/(CP.CF_MULBERRY_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=10,value=temp)  # 桑园

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=33).value or 0) / 1000)/(CP.CF_SEEDING_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=11,value=temp)  # 苗木

                temp = CP.LOSS_ORCHARD_TN * ((converted_sheet.cell(column=col, row=35).value or 0) / 1000)/(CP.CF_OTHER_MIN)
                TN_MIN_sheet.cell(column=col - start_col + 2, row=12,value=temp)  # 其他

        def TP_MAX_table_create():
            """
            总氮——MAX表创建
            """
            headers = ['', '乡镇1', '乡镇2', '乡镇3','合计']
            for col, header in enumerate(headers, start=1):
                TP_MAX_sheet.cell(row=1, column=col, value=header)

            TP_MAX_sheet.cell(1, 1, '乡镇')
            TP_MAX_sheet.cell(2, 1, '粮食')
            TP_MAX_sheet.cell(3, 1, '油料')
            TP_MAX_sheet.cell(4, 1, '中草药材')
            TP_MAX_sheet.cell(5, 1, '蔬菜')
            TP_MAX_sheet.cell(6, 1, '瓜果')
            TP_MAX_sheet.cell(7, 1, '茶园')
            TP_MAX_sheet.cell(8, 1, '花卉')
            TP_MAX_sheet.cell(9, 1, '林果园')
            TP_MAX_sheet.cell(10, 1, '桑园')
            TP_MAX_sheet.cell(11, 1, '苗木')
            TP_MAX_sheet.cell(12, 1, '其他')

            start_col = 3
            for col in range(start_col, converted_sheet.max_column + 1):
                temp = CP.LOSS_CROP_TP * ((converted_sheet.cell(column=col, row=12).value or 0) / 1000)/(CP.CF_FOOD_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=2,value=temp)  # 粮食

                temp = CP.LOSS_CROP_TP * ((converted_sheet.cell(column=col, row=14).value or 0) / 1000)/(CP.CF_OIL_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=3,value=temp)  # 油料

                temp = CP.LOSS_CROP_TP * ((converted_sheet.cell(column=col, row=16).value or 0) / 1000)/(CP.CF_MEDICINAL_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=4,value=temp)  # 中药材

                temp = CP.LOSS_CROP_TP * (((converted_sheet.cell(column=col, row=17).value or 0)/1000)/(CP.CF_VEGETABLE_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=18).value or 0)/ 1000) / (CP.CF_VEGETABLE_OPEN_MAX))
                TP_MAX_sheet.cell(column=col - start_col + 2, row=5,value=temp)  # 蔬菜类

                temp = CP.LOSS_ORCHARD_TP * (((converted_sheet.cell(column=col, row=20).value or 0)/1000)/(CP.CF_MELON_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=21).value or 0)/ 1000) / (CP.CF_MELON_OPEN_MAX))
                TP_MAX_sheet.cell(column=col - start_col + 2, row=6,value=temp)  # 瓜果类

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=24).value or 0) / 1000)/(CP.CF_TEA_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=7,value=temp)  # 茶园

                temp = CP.LOSS_ORCHARD_TP * (((converted_sheet.cell(column=col, row=25).value or 0)/1000)/(CP.CF_FLOWER_FACILITY_MAX) + \
                    ((converted_sheet.cell(column=col, row=26).value or 0)/ 1000) / (CP.CF_FLOWER_OPEN_MAX))
                TP_MAX_sheet.cell(column=col - start_col + 2, row=8,value=temp)  # 花卉类

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=29).value or 0) / 1000)/(CP.CF_FRUIT_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=9,value=temp)  # 林果园

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=31).value or 0) / 1000)/(CP.CF_MULBERRY_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=10,value=temp)  # 桑园

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=33).value or 0) / 1000)/(CP.CF_SEEDING_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=11,value=temp)  # 苗木

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=35).value or 0) / 1000)/(CP.CF_OTHER_MAX)
                TP_MAX_sheet.cell(column=col - start_col + 2, row=12,value=temp)  # 其他

        def TP_MIN_table_create():
            """
            总氮——MIN表创建
            """
            headers = ['', '乡镇1', '乡镇2', '乡镇3','合计']
            for col, header in enumerate(headers, start=1):
                TP_MIN_sheet.cell(row=1, column=col, value=header)

            TP_MIN_sheet.cell(1, 1, '乡镇')
            TP_MIN_sheet.cell(2, 1, '粮食')
            TP_MIN_sheet.cell(3, 1, '油料')
            TP_MIN_sheet.cell(4, 1, '中草药材')
            TP_MIN_sheet.cell(5, 1, '蔬菜')
            TP_MIN_sheet.cell(6, 1, '瓜果')
            TP_MIN_sheet.cell(7, 1, '茶园')
            TP_MIN_sheet.cell(8, 1, '花卉')
            TP_MIN_sheet.cell(9, 1, '林果园')
            TP_MIN_sheet.cell(10, 1, '桑园')
            TP_MIN_sheet.cell(11, 1, '苗木')
            TP_MIN_sheet.cell(12, 1, '其他')

            start_col = 3
            for col in range(start_col, converted_sheet.max_column + 1):
                temp = CP.LOSS_CROP_TP * ((converted_sheet.cell(column=col, row=12).value or 0) / 1000)/(CP.CF_FOOD_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=2,value=temp)  # 粮食

                temp = CP.LOSS_CROP_TP * ((converted_sheet.cell(column=col, row=14).value or 0) / 1000)/(CP.CF_OIL_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=3,value=temp)  # 油料

                temp = CP.LOSS_CROP_TP * ((converted_sheet.cell(column=col, row=16).value or 0) / 1000)/(CP.CF_MEDICINAL_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=4,value=temp)  # 中药材

                temp = CP.LOSS_CROP_TP * (((converted_sheet.cell(column=col, row=17).value or 0)/1000)/(CP.CF_VEGETABLE_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=18).value or 0)/ 1000) / (CP.CF_VEGETABLE_OPEN_MIN))
                TP_MIN_sheet.cell(column=col - start_col + 2, row=5,value=temp)  # 蔬菜类

                temp = CP.LOSS_ORCHARD_TP * (((converted_sheet.cell(column=col, row=20).value or 0)/1000)/(CP.CF_MELON_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=21).value or 0)/ 1000) / (CP.CF_MELON_OPEN_MIN))
                TP_MIN_sheet.cell(column=col - start_col + 2, row=6,value=temp)  # 瓜果类

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=24).value or 0) / 1000)/(CP.CF_TEA_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=7,value=temp)  # 茶园

                temp = CP.LOSS_ORCHARD_TP * (((converted_sheet.cell(column=col, row=25).value or 0)/1000)/(CP.CF_FLOWER_FACILITY_MIN) + \
                    ((converted_sheet.cell(column=col, row=26).value or 0)/ 1000) / (CP.CF_FLOWER_OPEN_MIN))
                TP_MIN_sheet.cell(column=col - start_col + 2, row=8,value=temp)  # 花卉类

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=29).value or 0) / 1000)/(CP.CF_FRUIT_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=9,value=temp)  # 林果园

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=31).value or 0) / 1000)/(CP.CF_MULBERRY_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=10,value=temp)  # 桑园

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=33).value or 0) / 1000)/(CP.CF_SEEDING_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=11,value=temp)  # 苗木

                temp = CP.LOSS_ORCHARD_TP * ((converted_sheet.cell(column=col, row=35).value or 0) / 1000)/(CP.CF_OTHER_MIN)
                TP_MIN_sheet.cell(column=col - start_col + 2, row=12,value=temp)  # 其他

        def SUM_table_create():
            headers = [' ', '氨氮', '总氮', '总磷']
            for col, header in enumerate(headers, start=1):
                SUM_sheet.cell(row=1, column=col, value=header)
            SUM_sheet.cell(row=2,column=1,value = "种植业排放营养物流失量（t）（最大复种指数）：")

            max_column = AN_MAX_sheet.max_column
            max_row = AN_MAX_sheet.max_row
            # 初始化累加器
            total = 0
            # 叠加最后一列的数据
            for row in range(2, max_row + 1):
                cell = AN_MAX_sheet.cell(row=row, column=max_column)
                # 假设单元格中是数字，如果是字符串需要转换
                if type(cell.value) is float:
                     value = cell.value
                total += value
            SUM_sheet.cell(row=2, column=2, value=total)

            max_column = TN_MAX_sheet.max_column
            max_row = TN_MAX_sheet.max_row
            total = 0
            for row in range(2, max_row + 1):
                cell = TN_MAX_sheet.cell(row=row, column=max_column)
                # 假设单元格中是数字，如果是字符串需要转换
                if type(cell.value) is float:
                    value = cell.value
                total += value
            SUM_sheet.cell(row=2, column=3, value=total)

            max_column = TP_MAX_sheet.max_column
            max_row = TP_MAX_sheet.max_row
            total = 0
            for row in range(2, max_row + 1):
                cell = TP_MAX_sheet.cell(row=row, column=max_column)
                if type(cell.value) is float:
                    value = cell.value
                total += value
            SUM_sheet.cell(row=2, column=4, value=total)

            max_column = AN_MIN_sheet.max_column
            max_row = AN_MIN_sheet.max_row
            # 初始化累加器
            total = 0
            # 叠加最后一列的数据
            for row in range(2, max_row + 1):
                cell = AN_MIN_sheet.cell(row=row, column=max_column)
                # 假设单元格中是数字，如果是字符串需要转换
                if type(cell.value) is float:
                     value = cell.value
                total += value
            SUM_sheet.cell(row=3, column=2, value=total)

            max_column = TN_MIN_sheet.max_column
            max_row = TN_MIN_sheet.max_row
            total = 0
            for row in range(2, max_row + 1):
                cell = TN_MIN_sheet.cell(row=row, column=max_column)
                # 假设单元格中是数字，如果是字符串需要转换
                if type(cell.value) is float:
                    value = cell.value
                total += value
            SUM_sheet.cell(row=3, column=3, value=total)

            max_column = TP_MIN_sheet.max_column
            max_row = TP_MIN_sheet.max_row
            total = 0
            for row in range(2, max_row + 1):
                cell = TP_MIN_sheet.cell(row=row, column=max_column)
                if type(cell.value) is float:
                    value = cell.value
                total += value
            SUM_sheet.cell(row=3, column=4, value=total)

        def beautify_sheet(sheet):
            """
            进行美化
            """
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            border_style = Side(style='thin', color='000000')
            border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = border
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 2
                sheet.column_dimensions[column].width = adjusted_width

        Mu_to_hec()
        Crop_area_statistics()
        AN_MAX_table_create()
        AN_MIN_table_create()
        TN_MAX_table_create()
        TN_MIN_table_create()
        TP_MAX_table_create()
        TP_MIN_table_create()
        SUM_table_create()

        beautify_sheet(crop_area_sheet)
        beautify_sheet(AN_MAX_sheet)
        beautify_sheet(AN_MIN_sheet)
        beautify_sheet(TN_MAX_sheet)
        beautify_sheet(TN_MIN_sheet)
        beautify_sheet(TP_MAX_sheet)
        beautify_sheet(TP_MIN_sheet)
        beautify_sheet(SUM_sheet)
        print(os.path.join(result_folder, f'result_{file_name}'))
        if not os.path.exists(result_folder):
            os.makedirs(result_folder)
        wb.save(os.path.join(result_folder, f'result_{file_name}'))
        print("down!")